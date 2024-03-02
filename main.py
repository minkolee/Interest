import openpyxl
import datetime

# 基础参数

# 银行每日计息的分割比例
base_tick = 1 / 360

# 日期间隔为1天
delta_date = datetime.timedelta(days=1)

# 设置利息总和变量
interest_sum = 0

# Ver 0.1 最简单的核心逻辑 - 从借据一开始放款计算到今天的累计利息

# 1 读出某个贷款项下的全部借据（不在程序里演示）
# 2 针对其中每一个借据，执行以下步骤
#   2.1 读入该借据的起始金额，起始日期，起始利率，设置变量：利息=0
#   2.2 读入该借据的本金变动序列和利率变动序列并判断长度，如果某一个序列等于0，设置一个对应的布尔值，无需进行更新
#       如果需要更新，设置指针指向序列的序号
#   2.3 设置三个变量：当前余额，当前日期，当前利率，并分别将其初始化为该借据的起始金额，起始日期，起始利率
#   2.4 开始循环，判断贷款的余额是否已经变为0和当前日期是否已经等于今天，满足任意一个条件退出循环，否则继续执行循环
#      循环体内部：
#
#             计算 当天利息 = 当前余额 * 当前利率 * base_tick
#             累加 当天利息 到循环外部的变量 利息总和 上：利息总和 = 利息总和 + 当天利息
#             将当前日期往后移动一天 当前日期 = 当前日期 + delta_date
#             根据本金变动序列和利率变动序列的布尔值确定是否要进行更新
#             不进行更新，继续执行循环
#             进行更新：
#                   判断日期是否等于当前序列指针指向的日期
#                      如果相等，将序列指针增加1（同时判断是否超过序列长度，超过则不再增加），然后将当前余额和当前利率更新为对应序列中新的金额
#                      如果不相等，继续执行循环
#   2.5 跳出循环后，利息变量中存储的是该贷款从放款日至今天的累计利息

# 以sample.xlsx中note表中id为1的借据演示代码

# 打开sample.xlsx工作簿及三张工作表
wb = openpyxl.open("sample.xlsx")
note = wb["note"]
balance_series = wb["balance_series"]
ir_series = wb["ir_series"]

# 当前借据的id，这里写死，实际在web开发中应该通过路径从数据库中获取
current_note_id = note.cell(2, 1).value

# 读入借据01的起始金额，起始日期，起始利率
start_balance = note.cell(2, 2).value
start_date = note.cell(2, 3).value.date()
start_ir = note.cell(2, 4).value
print(
    "借据id为 {} 的借据起始金额是 {:.2f} ，起始日期是 {} ，起始利率是 {:.4%} "
    "\n--------------------------------------------------------------------------------".format(
        current_note_id, start_balance, start_date, start_ir))


# 读入该借据对应的本金变动序列和利率变动序列
# 由于本金变动序列和利率变动序列的数据前四列相同，可以采用同一个函数读取，web开发中应采取根据外键查询的方式读取
def load_series(note_id, worksheet):
    name = worksheet.cell(1, 3).value
    print("加载{}序列".format(name))
    series = []
    max_row = worksheet.max_row
    for i in range(2, max_row + 1):
        if worksheet.cell(i, 2).value == note_id:
            series.append((worksheet.cell(i, 3).value, worksheet.cell(i, 4).value.date()))
    show_series(series, name)
    print("--------------------------------------------------------------------------------")
    return series


# 显示读入的序列信息的函数
def show_series(series, name):
    if len(series) == 0:
        print("序列为空")
    else:
        print("序列信息：")
        if name == "借据余额":
            for i in series:
                print("{}：{:.2f} \t变动日期：{}".format(name, i[0], i[1]))
        else:
            for i in series:
                print("{}：{:.4%} \t变动日期：{}".format(name, i[0], i[1]))


# 通过函数读取和展示读入到的结果
note_balance_series = load_series(current_note_id, balance_series)
note_ir_series = load_series(current_note_id, ir_series)

# 判断本金序列和利率序列是否需要更新
is_balance_update = len(note_balance_series) != 0
is_ir_update = len(note_ir_series) != 0

# 在需要更新的情况下，设置序列的指针，实际上不更新不需要设置，这里为了简便都设置为0,在后续代码中先判断布尔值
balance_series_pointer = 0
ir_series_pointer = 0

# 核心逻辑

# 获取今天的日期
end_date = datetime.date.today()
print("开始计算截至到今天({})的利息".format(end_date))

# 设置当前金额，当前利率，当前日期为初始三变量
current_balance = start_balance
current_date = start_date
current_ir = start_ir
print(
    "当前金额是 {:.2f} ，起始日期是 {} ，起始利率是 {:.4%} \n--------------------------------------------------------------------------------".format(
        current_balance, current_date, current_ir))

print("本金序列是否需要更新：{}\t利息序列是否需要更新：{}".format(is_balance_update, is_ir_update))

# 启动核心逻辑循环

# 当前日期不等于今天并同时满足本金还不为0的情况下，执行循环
while current_date != end_date and current_balance != 0:
    print("当前本金为：{:.2f}\t\t当前利率为：{:.4%}\t\t当前日期为：{}\t\t当前累计利息为：{:.2f}".format(current_balance, current_ir, current_date,
                                                                              interest_sum))
    # 计算当天的利息并累加到利息总和上
    interest_sum = interest_sum + current_ir * current_balance * base_tick
    # 让日期前进一天
    current_date = current_date + delta_date

    # 判断本金序列是否需要更新
    if is_balance_update:
        # 比较当前日期与本金序列指针指向的日期
        if current_date == note_balance_series[balance_series_pointer][1]:
            # 如果相同，更新当前剩余本金金额
            current_balance = note_balance_series[balance_series_pointer][0]
            # 更新成功后移动指针，如果指针已经到最后，不需要移动
            if balance_series_pointer < len(note_balance_series) - 1:
                balance_series_pointer = balance_series_pointer + 1

    # 判断利率序列是否需要更新
    if is_ir_update:
        # 比较当前日期与本金序列指针指向的日期
        if current_date == note_ir_series[ir_series_pointer][1]:
            # 如果相同，更新当前剩余本金金额
            current_ir = note_ir_series[ir_series_pointer][0]
            # 更新成功后移动指针，如果指针已经到最后，不需要移动
            if ir_series_pointer < len(note_ir_series) - 1:
                ir_series_pointer = ir_series_pointer + 1

print("当前本金为：{:.2f}\t\t当前利率为：{:.4%}\t\t当前日期为：{}\t\t当前累计利息为：{:.2f}".format(current_balance, current_ir, current_date,
                                                                          interest_sum))

input()
